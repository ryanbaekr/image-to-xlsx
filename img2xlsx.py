from openpyxl import Workbook
from openpyxl.styles import PatternFill
from PIL import Image
import glob

def toAlpha(dec):
    chars="ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if (dec >= 0):
        x=dec%26
        rest=dec//26
        if (rest == 0):
            return chars[x]
        else:
            return toAlpha(rest-1)+chars[x]

wb=Workbook()
ws=wb.active

im=Image.open(glob.glob("./img/*")[0])
col,row=im.size
pixels=im.load()

for i in range(col):
    column=toAlpha(i)
    ws.column_dimensions[column].width=1
    for j in range(row):
        if (i == 0):
            ws.row_dimensions[j].height=5.5
        if (len(pixels[i,j]) == 3):
            r,g,b=pixels[i,j]
        else:
            r,g,b,a=pixels[i,j]
        cellName=column+str(j+1)
        fillColor=str((hex(r//8*8*256*256+g//8*8*256+b//8*8).split('x')[-1]).zfill(6))
        cell=ws[cellName]
        fill=PatternFill("solid",fgColor=fillColor)
        cell.fill=fill

wb.save("xlsx/output.xlsx")
